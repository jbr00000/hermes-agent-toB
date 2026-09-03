"""从 BULL-cn 演示库生成 NL2SQL 元数据冷启动 Excel（每库一份，7-sheet 与导入模板同构）。

数据来源：
  - 表结构 sheet：直连 MySQL 逐表 SHOW CREATE TABLE（说明取 db_info.json 的表中文名，
    MySQL 里这些表没有 table_comment）；provider=AI
  - 外键关系 sheet：db_info.json 的 table_rel；provider=AI
  - 范例 sheet：dev.json 按 db_name 分库（question + sql_query）；provider=AI
  - 维度 sheet：对低基数 varchar/char 列 SELECT DISTINCT 生成码表草稿；provider=AI，
    备注标「自动扫描草稿，待人工确认」
  - 术语 / 指标 sheet：只写表头留空——这两类必须人工补口径，机器生成没有可信度

连接密码从 NL2SQL_SRC_PASSWORD 环境变量读取，不打印、不落盘。

用法：
    NL2SQL_SRC_PASSWORD=... python scripts/build_nl2sql_metadata_xlsx.py \
        [--mysql-host 127.0.0.1] [--mysql-port 13306] [--mysql-user root] \
        [--bull-dir nl2sql_data/BULL/BULL/BULL-cn] [--out-dir nl2sql_data/metadata_xlsx]
"""
from __future__ import annotations

import argparse
import io
import json
import os
import sys
from pathlib import Path

import pymysql
from openpyxl import Workbook

# ---- 与 server/nl2sql/excel.py 的 _SHEETS 保持同构（脚本独立，不 import server 包，
# ---- 避免拉起 storage 引擎；改 excel.py 布局时必须同步这里）
_SHEETS: dict[str, list[str]] = {
    "表结构": ["提供者", "表名", "建表语句", "说明"],
    "术语": ["提供者", "术语名", "术语解释", "近义词", "备注"],
    "指标": ["提供者", "指标名", "指标展示名", "计算口径", "备注"],
    "维度": ["提供者", "维度名", "维度展示名", "标准值Key", "标准值", "备注"],
    "范例": ["提供者", "问题", "参考SQL", "备注"],
    "外键关系": ["提供者", "源表格", "源表字段", "目标表", "目标表字段", "关联说明"],
}

# BULL-cn 库名 → 本仓 MySQL 里的演示库名
_DB_NAME_MAP = {
    "ccks_fund": "nl2sql_fund",
    "ccks_stock": "nl2sql_stock",
    "ccks_macro": "nl2sql_macro",
}

_MAX_TEXT = 32767  # Excel 单元格上限，DDL 不会触到，防御性截断


def _cell(value: object) -> str:
    text = "" if value is None else str(value)
    return text[:_MAX_TEXT]


def _distinct_candidates(conn, database: str) -> list[tuple[str, str]]:
    """库里所有文本列（低基数维度的候选池）。

    BULL 库的字符串列是 text/tinytext（character_maximum_length 为 NULL），
    不能按声明长度过滤——值长度在 fetch 后用 max_value_len 把关。
    """
    with conn.cursor() as cur:
        cur.execute(
            "SELECT table_name, column_name FROM information_schema.columns "
            "WHERE table_schema=%s AND data_type IN ('varchar','char','enum','text','tinytext') "
            "ORDER BY table_name, ordinal_position",
            (database,),
        )
        return [(str(t), str(c)) for t, c in cur.fetchall()]


def _scan_dimensions(
    conn, database: str, chi_names: dict[tuple[str, str], str],
    min_card: int, max_card: int, max_value_len: int,
) -> tuple[list[list[str]], int]:
    """低基数字段 SELECT DISTINCT → 维度码表草稿。返回 (rows, 扫描过的候选列数)。"""
    rows: list[list[str]] = []
    seen: set[tuple[str, str]] = set()
    candidates = _distinct_candidates(conn, database)
    with conn.cursor() as cur:
        for table, column in candidates:
            cur.execute(
                f"SELECT COUNT(DISTINCT `{column}`) FROM `{database}`.`{table}`"
            )
            count = int(cur.fetchone()[0] or 0)
            if not (min_card <= count <= max_card):
                continue
            cur.execute(
                f"SELECT DISTINCT `{column}` FROM `{database}`.`{table}` "
                f"WHERE `{column}` IS NOT NULL ORDER BY `{column}` LIMIT %s",
                (max_card + 1,),
            )
            values = [str(v[0]).strip() for v in cur.fetchall()]
            values = [v for v in values if v and len(v) <= max_value_len]
            if len(values) < min_card:
                continue
            display = chi_names.get((table, column), "")
            dim_name = f"{table}.{column}"
            for value in values:
                if (dim_name, value) in seen:
                    continue
                seen.add((dim_name, value))
                rows.append(["AI", dim_name, display, value, value, "自动扫描草稿，待人工确认"])
    return rows, len(candidates)


def build_one(
    conn, bull_dir: Path, entry: dict, dev_rows: list[dict],
    min_card: int, max_card: int, max_value_len: int,
) -> tuple[bytes, dict[str, int]]:
    """单库生成一份 xlsx，返回 (bytes, 各 sheet 行数)。"""
    ccks_name = entry["db_name"]
    database = _DB_NAME_MAP[ccks_name]
    zh_tables = {en: zh for en, zh in entry["table_name"]}
    chi_names: dict[tuple[str, str], str] = {}
    for table_info in entry["column_info"]:
        table = table_info["table"]
        for col, chi in zip(table_info["columns"][1:], table_info["column_chiName"][1:]):
            if chi:
                chi_names[(table, col)] = str(chi)

    wb = Workbook()
    intro = wb.active
    intro.title = "填写说明"
    intro.append(["适用数据集", database])
    intro.append(["文件格式", ".xlsx，一个元数据类型一个 sheet；首行表头，首列为提供者（人工/AI）"])
    intro.append(["导入规则", "按各 sheet 的关键字段去重：已存在且内容一致跳过、内容不同更新、不存在新增"])
    intro.append(["填写说明 sheet", "仅说明用途，导入时忽略"])
    intro.append(["来源", f"BULL-cn {ccks_name} 冷启动：表结构=SHOW CREATE TABLE，外键=db_info.json，"
                          "范例=dev.json，维度=低基数字段自动扫描（待人工确认）；术语/指标需人工补"])

    counts: dict[str, int] = {}

    # 表结构
    ws = wb.create_sheet("表结构")
    ws.append(_SHEETS["表结构"])
    with conn.cursor() as cur:
        for en, _zh in entry["table_name"]:
            cur.execute(f"SHOW CREATE TABLE `{database}`.`{en}`")
            row = cur.fetchone()
            if row is None:
                continue
            ws.append(["AI", en, _cell(row[1]), zh_tables.get(en, "")])
    counts["表结构"] = ws.max_row - 1

    # 术语 / 指标：留空待人工补
    for name in ("术语", "指标"):
        ws = wb.create_sheet(name)
        ws.append(_SHEETS[name])
        counts[name] = 0

    # 维度
    dim_rows, scanned = _scan_dimensions(
        conn, database, chi_names, min_card, max_card, max_value_len
    )
    ws = wb.create_sheet("维度")
    ws.append(_SHEETS["维度"])
    for row in dim_rows:
        ws.append([_cell(v) for v in row])
    counts["维度"] = len(dim_rows)
    counts["维度候选列"] = scanned

    # 范例
    ws = wb.create_sheet("范例")
    ws.append(_SHEETS["范例"])
    for item in dev_rows:
        ws.append(["AI", _cell(item["question"]), _cell(item["sql_query"]), "BULL-cn dev.json"])
    counts["范例"] = len(dev_rows)

    # 外键关系
    ws = wb.create_sheet("外键关系")
    ws.append(_SHEETS["外键关系"])
    seen_fk: set[tuple[str, str, str, str]] = set()
    fk_count = 0
    for rel in entry["table_rel"]:
        (src_table, src_col), (tgt_table, tgt_col) = rel[0], rel[1]
        key = (src_table, src_col, tgt_table, tgt_col)
        if key in seen_fk:
            continue
        seen_fk.add(key)
        ws.append(["AI", src_table, src_col, tgt_table, tgt_col, ""])
        fk_count += 1
    counts["外键关系"] = fk_count

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), counts


def main() -> int:
    parser = argparse.ArgumentParser(description="生成 BULL 三库的 NL2SQL 元数据冷启动 Excel")
    parser.add_argument("--mysql-host", default="127.0.0.1")
    parser.add_argument("--mysql-port", type=int, default=13306)
    parser.add_argument("--mysql-user", default="root")
    parser.add_argument("--bull-dir", default="nl2sql_data/BULL/BULL/BULL-cn")
    parser.add_argument("--out-dir", default="nl2sql_data/metadata_xlsx")
    parser.add_argument("--min-card", type=int, default=2, help="维度候选：distinct 值下限")
    parser.add_argument("--max-card", type=int, default=30, help="维度候选：distinct 值上限")
    parser.add_argument("--max-value-len", type=int, default=50, help="维度标准值长度上限")
    args = parser.parse_args()

    password = os.environ.get("NL2SQL_SRC_PASSWORD")
    if not password:
        print("ERROR: 需要 NL2SQL_SRC_PASSWORD 环境变量（MySQL 连接密码）", file=sys.stderr)
        return 2

    bull_dir = Path(args.bull_dir)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    db_info = json.loads((bull_dir / "db_info.json").read_text(encoding="utf-8"))
    dev = json.loads((bull_dir / "dev.json").read_text(encoding="utf-8"))
    dev_by_db: dict[str, list[dict]] = {}
    for item in dev:
        dev_by_db.setdefault(item["db_name"], []).append(item)

    conn = pymysql.connect(
        host=args.mysql_host, port=args.mysql_port, user=args.mysql_user,
        password=password, connect_timeout=5, read_timeout=60,
    )
    try:
        for entry in db_info:
            ccks_name = entry["db_name"]
            if ccks_name not in _DB_NAME_MAP:
                print(f"跳过未知库 {ccks_name}")
                continue
            database = _DB_NAME_MAP[ccks_name]
            content, counts = build_one(
                conn, bull_dir, entry, dev_by_db.get(ccks_name, []),
                args.min_card, args.max_card, args.max_value_len,
            )
            out_path = out_dir / f"元数据配置-{database}.xlsx"
            out_path.write_bytes(content)
            detail = "，".join(f"{k} {v}" for k, v in counts.items() if k != "维度候选列")
            print(f"[{database}] {out_path} — {detail}（维度扫描候选列 {counts['维度候选列']}）")
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
