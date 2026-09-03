"""NL2SQL 元数据 Excel：模板下载 / 导入解析（preview→confirm）/ 导出。

文件格式与 lone-ai「元数据配置导出数据」模板严格同构——1 个填写说明
sheet + 6 个资产 sheet（一类元数据一个 sheet，首列均为「提供者」），
保证 lone-ai 侧导出的历史文件可直接导入本系统，本系统模板也可反向使用。

  表结构:  提供者 / 表名 / 建表语句 / 说明
  术语:    提供者 / 术语名 / 术语解释 / 近义词 / 备注
  指标:    提供者 / 指标名 / 指标展示名 / 计算口径 / 备注
  维度:    提供者 / 维度名 / 维度展示名 / 标准值Key / 标准值 / 备注
  范例:    提供者 / 问题 / 参考SQL / 备注
  外键关系: 提供者 / 源表格 / 源表字段 / 目标表 / 目标表字段 / 关联说明
"""
from __future__ import annotations

import io
import threading
import time
import uuid
from typing import Any

from openpyxl import Workbook, load_workbook

from server.nl2sql import store

# sheet 名 → (kind, 表头, 行 tuple → fields dict 的键序)
_SHEETS: dict[str, tuple[str, list[str], list[str]]] = {
    "表结构": ("tables", ["提供者", "表名", "建表语句", "说明"],
               ["provider", "table_name", "ddl_content", "description"]),
    "术语": ("terms", ["提供者", "术语名", "术语解释", "近义词", "备注"],
             ["provider", "terminology", "terminology_explain", "synonyms", "remark"]),
    "指标": ("metrics", ["提供者", "指标名", "指标展示名", "计算口径", "备注"],
             ["provider", "index_name", "index_display_name", "calculate_method", "remark"]),
    "维度": ("dimensions", ["提供者", "维度名", "维度展示名", "标准值Key", "标准值", "备注"],
             ["provider", "dimension_name", "dimension_display_name", "db_data_key", "db_data_value", "remark"]),
    "范例": ("examples", ["提供者", "问题", "参考SQL", "备注"],
             ["provider", "question", "question_sql", "remark"]),
    "外键关系": ("foreignKeys", ["提供者", "源表格", "源表字段", "目标表", "目标表字段", "关联说明"],
                 ["provider", "source_table", "source_column", "target_table", "target_column", "relation_desc"]),
}

# 各 kind 的必填字段（缺任一 → 行级 error）
_REQUIRED: dict[str, tuple[str, ...]] = {
    "tables": ("table_name", "ddl_content"),
    "terms": ("terminology",),
    "metrics": ("index_name",),
    "dimensions": ("dimension_name", "db_data_key", "db_data_value"),
    "foreignKeys": ("source_table", "source_column", "target_table", "target_column"),
    "examples": ("question", "question_sql"),
}

# 预览去重键：与既有记录比对判断 create / update / duplicate
_DEDUP_KEYS: dict[str, tuple[str, ...]] = {
    "tables": ("table_name",),
    "terms": ("terminology",),
    "metrics": ("index_name",),
    "dimensions": ("dimension_name", "db_data_key"),
    "foreignKeys": ("source_table", "source_column", "target_table", "target_column"),
    "examples": ("question",),
}

_PROVIDER_TO_EXCEL = {"MANUAL": "人工", "AI": "AI"}
_PROVIDER_FROM_EXCEL = {"人工": "MANUAL", "AI": "AI"}

_PREVIEW_TTL_SECONDS = 30 * 60
_PREVIEWS: dict[str, dict[str, Any]] = {}
_PREVIEWS_LOCK = threading.Lock()


def _to_excel_provider(provider: str) -> str:
    return _PROVIDER_TO_EXCEL.get(provider, "人工")


def _from_excel_provider(value: str) -> str:
    return _PROVIDER_FROM_EXCEL.get(value.strip(), "MANUAL")


def _write_workbook(bundle: dict[str, list[dict[str, Any]]], dataset_name: str) -> bytes:
    wb = Workbook()
    intro = wb.active
    intro.title = "填写说明"
    intro.append(["适用数据集", dataset_name])
    intro.append(["文件格式", ".xlsx，一个元数据类型一个 sheet；首行表头，首列为提供者（人工/AI）"])
    intro.append(["导入规则", "按各 sheet 的关键字段去重：已存在且内容一致跳过、内容不同更新、不存在新增"])
    intro.append(["填写说明 sheet", "仅说明用途，导入时忽略"])
    for sheet_name, (kind, headers, keys) in _SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for row in bundle.get(kind, []):
            ws.append([
                _to_excel_provider(str(row.get("provider") or "MANUAL")),
                *[_cell_text(row.get(key)) for key in keys[1:]],
            ])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    return str(value)


def build_template(dataset_name: str) -> bytes:
    """空模板（只有表头+填写说明）。"""
    return _write_workbook({}, dataset_name)


def build_export(dataset_name: str, bundle: dict[str, list[dict[str, Any]]]) -> bytes:
    return _write_workbook(bundle, dataset_name)


def _parse_rows(content: bytes) -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]], list[str]]:
    """解析 xlsx → (按 kind 分组的 fields 列表, 行级错误, 忽略的 sheet 名)。"""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parsed: dict[str, list[dict[str, Any]]] = {kind: [] for _, (kind, _, _) in _SHEETS.items()}
    errors: list[dict[str, Any]] = []
    ignored: list[str] = []
    for sheet_name in wb.sheetnames:
        if sheet_name == "填写说明":
            ignored.append(sheet_name)
            continue
        spec = _SHEETS.get(sheet_name)
        if spec is None:
            ignored.append(sheet_name)
            continue
        kind, _headers, keys = spec
        ws = wb[sheet_name]
        for row_index, cells in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            values = ["" if cell is None else str(cell).strip() for cell in cells]
            if not any(values):
                continue
            fields: dict[str, Any] = {}
            for position, key in enumerate(keys):
                if position >= len(values):
                    break
                if key == "provider":
                    fields[key] = _from_excel_provider(values[position])
                else:
                    fields[key] = values[position]
            missing = [key for key in _REQUIRED[kind] if not fields.get(key)]
            if missing:
                errors.append({
                    "sheet": sheet_name, "row": row_index,
                    "message": f"缺少必填列（{', '.join(missing)}）",
                })
                continue
            parsed[kind].append(fields)
    wb.close()
    return parsed, errors, ignored


def _dedup_key(kind: str, fields: dict[str, Any]) -> tuple:
    return tuple(str(fields.get(key) or "").strip() for key in _DEDUP_KEYS[kind])


def create_preview(dataset_id: str, content: bytes) -> dict[str, Any]:
    """解析 + 与既有元数据比对，返回预览摘要；待确认项挂到 previewId（30 分钟有效）。"""
    parsed, errors, ignored = _parse_rows(content)
    existing = store.get_meta_bundle(dataset_id)

    type_summaries: dict[str, dict[str, int]] = {}
    plan: dict[str, list[dict[str, Any]]] = {}
    for kind, rows in parsed.items():
        summary = {"read": len(rows), "create": 0, "update": 0, "duplicate": 0, "error": 0}
        existing_by_key = {_dedup_key(kind, row): row for row in existing.get(kind, [])}
        actions: list[dict[str, Any]] = []
        for fields in rows:
            hit = existing_by_key.get(_dedup_key(kind, fields))
            if hit is None:
                summary["create"] += 1
                actions.append({"action": "create", "fields": fields})
                continue
            # 内容比对只覆盖 Excel 里有的列；enabled 不在 Excel 中，
            # 导入更新不触碰既有行的启用状态
            changed = any(
                str(hit.get(key) or "") != str(value or "")
                for key, value in fields.items()
            )
            if changed:
                summary["update"] += 1
                actions.append({"action": "update", "id": hit["id"], "fields": fields})
            else:
                summary["duplicate"] += 1
        summary["error"] = sum(1 for e in errors if _SHEETS.get(e["sheet"], (None,))[0] == kind)
        type_summaries[kind] = summary
        plan[kind] = actions

    preview_id = uuid.uuid4().hex
    with _PREVIEWS_LOCK:
        # 顺手清理过期预览，防止内存只增不减
        expired = [pid for pid, p in _PREVIEWS.items() if p["expires_at"] < time.time()]
        for pid in expired:
            del _PREVIEWS[pid]
        _PREVIEWS[preview_id] = {
            "dataset_id": dataset_id,
            "plan": plan,
            "expires_at": time.time() + _PREVIEW_TTL_SECONDS,
        }
    return {
        "preview_id": preview_id,
        "type_summaries": type_summaries,
        "errors": errors[:200],  # 行级错误截断，防超长响应
        "ignored_sheets": ignored,
    }


def confirm_preview(dataset_id: str, preview_id: str) -> dict[str, int]:
    """按预览计划落库（create + update）。返回 {created, updated}。"""
    with _PREVIEWS_LOCK:
        preview = _PREVIEWS.pop(preview_id, None)
    if preview is None or preview["expires_at"] < time.time():
        raise KeyError(preview_id)
    if preview["dataset_id"] != dataset_id:
        raise ValueError("preview 不属于该数据集")

    created = updated = 0
    for kind, actions in preview["plan"].items():
        for action in actions:
            if action["action"] == "create":
                store.upsert_meta_item(dataset_id, kind, None, action["fields"])
                created += 1
            else:
                store.upsert_meta_item(dataset_id, kind, action["id"], action["fields"])
                updated += 1
    return {"created": created, "updated": updated}
